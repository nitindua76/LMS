"""
Professional-styled Excel export for admin course analytics — the first
thing in this stack that renders Excel, via openpyxl.

Three sheets: 'Summary' (course-level counts), 'Employee Progress' (one row
per targeted employee, with both best and latest quiz score plus attempts
used), and 'Quiz Attempts' (one row per individual attempt — the full audit
trail). Deliberately all three rather than picking one view of "the score":
compliance wants "did they eventually pass" (best), a manager coaching
someone wants "are they improving" (latest vs. first), and an audit wants
the complete history (every attempt) — these are different questions, and
a spreadsheet is cheap enough to just answer all of them at once.
"""
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.models.course import Course, Section, Quiz
from app.models.enrollment import (
    Enrollment, EnrollmentStatus, QuizAttempt, QuizAttemptStatus, SectionProgress,
)
from app.models.user import User
from app.routers.admin.analytics import get_targeted_user_ids, calculate_progress_pct

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F2937")
SUBTITLE_FONT = Font(size=10, color="6B7280", italic=True)
PASS_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
THIN_BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))


def _style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_course_report(db: Session, course: Course) -> BytesIO:
    wb = Workbook()
    _build_summary_sheet(db, course, wb.active)
    _build_progress_sheet(db, course, wb.create_sheet("Employee Progress"))
    _build_attempts_sheet(db, course, wb.create_sheet("Quiz Attempts"))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_summary_sheet(db: Session, course: Course, ws: Worksheet) -> None:
    ws.title = "Summary"
    targeted_ids = get_targeted_user_ids(db, course.id)
    enrollments = db.query(Enrollment).filter(Enrollment.course_id == course.id).all()
    completed = sum(1 for e in enrollments if e.status == EnrollmentStatus.completed)
    in_progress = sum(1 for e in enrollments if e.status == EnrollmentStatus.in_progress)
    failed = sum(1 for e in enrollments if e.status == EnrollmentStatus.failed)
    expired = sum(1 for e in enrollments if e.status == EnrollmentStatus.expired)
    not_enrolled = max(0, len(targeted_ids) - len(enrollments))

    ws["A1"] = course.title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Course report — generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = SUBTITLE_FONT

    rows = [
        ("Status", course.status.value),
        ("Mandatory", "Yes" if course.mandatory else "No"),
        ("Target audience size", len(targeted_ids)),
        ("Completed", completed),
        ("In progress", in_progress),
        ("Failed", failed),
        ("Expired", expired),
        ("Not yet enrolled", not_enrolled),
    ]
    for i, (label, value) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    _autosize_columns(ws, [24, 20])


def _build_progress_sheet(db: Session, course: Course, ws: Worksheet) -> None:
    targeted_ids = get_targeted_user_ids(db, course.id)
    users = (
        {u.id: u for u in db.query(User).options(joinedload(User.discipline), joinedload(User.level))
            .filter(User.id.in_(targeted_ids)).all()}
        if targeted_ids else {}
    )
    enrollments = {
        e.user_id: e for e in db.query(Enrollment).filter(Enrollment.course_id == course.id).all()
    }

    headers = ["Employee", "Email", "Department", "Level", "Status", "Progress %",
               "Best Score %", "Latest Score %", "Attempts Used", "Started", "Completed"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    row = 2
    for user_id in sorted(targeted_ids, key=lambda uid: users[uid].name if uid in users else ""):
        user = users.get(user_id)
        if not user:
            continue
        e = enrollments.get(user_id)

        attempts = []
        if e:
            attempts = (
                db.query(QuizAttempt)
                .filter(QuizAttempt.enrollment_id == e.id, QuizAttempt.status == QuizAttemptStatus.submitted)
                .order_by(QuizAttempt.attempt_no)
                .all()
            )
        best_score = max((a.score_pct for a in attempts if a.score_pct is not None), default=None)
        latest_score = attempts[-1].score_pct if attempts else None

        progress_pct = calculate_progress_pct(db, e.id, user_id, course.id) if e else 0
        completed_at = None
        if e and e.status == EnrollmentStatus.completed:
            completed_at = db.query(func.max(SectionProgress.completed_at)).filter(
                SectionProgress.enrollment_id == e.id
            ).scalar()

        values = [
            user.name, user.email,
            user.discipline.name if user.discipline else "—",
            user.level.name if user.level else "—",
            e.status.value if e else "not_started",
            progress_pct,
            best_score if best_score is not None else "—",
            latest_score if latest_score is not None else "—",
            len(attempts),
            e.started_at.strftime("%Y-%m-%d") if e and e.started_at else "—",
            completed_at.strftime("%Y-%m-%d") if completed_at else "—",
        ]
        for col, v in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=v).border = THIN_BORDER
        row += 1

    _autosize_columns(ws, [22, 28, 16, 12, 14, 12, 12, 14, 12, 12, 12])


def _build_attempts_sheet(db: Session, course: Course, ws: Worksheet) -> None:
    headers = ["Employee", "Email", "Section", "Attempt #", "Score %", "Result", "Submitted At"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    attempts = (
        db.query(QuizAttempt)
        .join(Quiz, Quiz.id == QuizAttempt.quiz_id)
        .join(Section, Section.id == Quiz.section_id)
        .join(Enrollment, Enrollment.id == QuizAttempt.enrollment_id)
        .join(User, User.id == Enrollment.user_id)
        .options(
            joinedload(QuizAttempt.quiz).joinedload(Quiz.section),
            joinedload(QuizAttempt.enrollment).joinedload(Enrollment.user),
        )
        .filter(Section.course_id == course.id, QuizAttempt.status == QuizAttemptStatus.submitted)
        .order_by(User.name, QuizAttempt.attempt_no)
        .all()
    )

    row = 2
    for a in attempts:
        values = [
            a.enrollment.user.name, a.enrollment.user.email, a.quiz.section.title, a.attempt_no,
            a.score_pct if a.score_pct is not None else "—",
            "Passed" if a.passed else "Failed",
            a.submitted_at.strftime("%Y-%m-%d %H:%M") if a.submitted_at else "—",
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.fill = PASS_FILL if a.passed else FAIL_FILL
            cell.border = THIN_BORDER
        row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="No quiz attempts recorded yet.")

    _autosize_columns(ws, [22, 28, 24, 10, 10, 10, 18])
